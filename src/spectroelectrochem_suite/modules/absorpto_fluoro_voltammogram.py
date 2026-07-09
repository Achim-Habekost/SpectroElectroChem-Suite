"""
Absorpto-/Fluorovoltammogramm GUI

CSV structure:
- First row: potentials
- First column: wavelengths
- Remaining cells: absorption / fluorescence intensity values

Outputs:
- Excel file with matrix, long-format data, waterfall data and embedded waterfall plot
- interactive HTML 3D surface
- interactive HTML heatmap
- interactive HTML contour plot
- rotatable HTML waterfall plot
"""

from pathlib import Path
import sys
import subprocess
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.ticker import MultipleLocator
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font


def ensure_package(import_name, pip_name=None):
    if pip_name is None:
        pip_name = import_name
    try:
        return __import__(import_name)
    except ModuleNotFoundError:
        answer = messagebox.askyesno(
            "Missing package",
            f"The package '{pip_name}' is missing.\n\nInstall it now?"
        )
        if answer:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
            return __import__(import_name)
        raise SystemExit(f"Missing package: {pip_name}")


def ensure_plotly():
    ensure_package("plotly", "plotly")
    import plotly.graph_objects as go
    return go


def safe_float(value):
    if pd.isna(value):
        return np.nan
    text = str(value).strip().replace(",", ".")
    if text == "":
        return np.nan
    return pd.to_numeric(text, errors="coerce")


def safe_name_number(x):
    return re.sub(r"[^0-9A-Za-z_.-]+", "_", f"{x:g}")


def read_voltammogram_csv(path: Path):
    encodings = ["utf-8-sig", "cp1252", "latin1", "utf-16", "utf-16le", "utf-16be"]
    last_error = None

    for enc in encodings:
        try:
            df = pd.read_csv(path, sep=None, engine="python", header=None, dtype=str, encoding=enc)
            encoding_used = enc
            break
        except UnicodeDecodeError as err:
            last_error = err
    else:
        raise UnicodeDecodeError("unknown", b"", 0, 1, "Could not decode CSV file.") from last_error

    df = df.map(safe_float)
    df = df.dropna(how="all").dropna(axis=1, how="all")

    potentials = df.iloc[0, 1:].to_numpy(dtype=float)
    wavelengths = df.iloc[1:, 0].to_numpy(dtype=float)
    values = df.iloc[1:, 1:].to_numpy(dtype=float)

    valid_pot = ~np.isnan(potentials)
    valid_wl = ~np.isnan(wavelengths)

    potentials = potentials[valid_pot]
    wavelengths = wavelengths[valid_wl]
    values = values[valid_wl, :][:, valid_pot]

    # Sort axes
    pot_order = np.argsort(potentials)
    potentials = potentials[pot_order]
    values = values[:, pot_order]

    wl_order = np.argsort(wavelengths)
    wavelengths = wavelengths[wl_order]
    values = values[wl_order, :]

    return potentials, wavelengths, values, encoding_used


def moving_average(values, window=5):
    if window < 3:
        return values.copy()
    if window % 2 == 0:
        window += 1
    kernel = np.ones(window) / window
    result = np.empty_like(values, dtype=float)
    for j in range(values.shape[1]):
        result[:, j] = np.convolve(values[:, j], kernel, mode="same")
    return result


def savgol_smooth(values, window=11, poly=3):
    ensure_package("scipy", "scipy")
    from scipy.signal import savgol_filter
    n = values.shape[0]
    window = int(window)
    poly = int(poly)
    if window % 2 == 0:
        window += 1
    if window > n:
        window = n if n % 2 == 1 else n - 1
    if window <= poly:
        window = poly + 2
        if window % 2 == 0:
            window += 1
    if window < 5 or window > n or window <= poly:
        return values.copy()
    return savgol_filter(values, window_length=window, polyorder=poly, axis=0, mode="interp")


def scale_values(values, mode):
    z = values.astype(float).copy()
    label = "Intensity / a.u."
    suffix = "raw"

    if mode in {"clip99", "clip99_log1p"}:
        lo = np.nanpercentile(z, 1)
        hi = np.nanpercentile(z, 99)
        z = np.clip(z, lo, hi)
        label = "Intensity clipped to 1-99 percentile"
        suffix = "clip99"

    if mode in {"log1p", "clip99_log1p"}:
        zmin = np.nanmin(z)
        if zmin < 0:
            z = z - zmin
        z = np.log1p(z)
        label = "log1p intensity" if mode == "log1p" else "clipped + log1p intensity"
        suffix = "log1p" if mode == "log1p" else "clip99_log1p"

    return z, label, suffix


def axis_settings_wavelength(start, final):
    tick0 = 50 * np.floor(min(start, final) / 50)
    return dict(title="Wavelength / nm", tickmode="linear", tick0=tick0, dtick=50)


def axis_settings_intensity(label, value_range):
    settings = dict(title=label)
    if value_range is not None:
        settings["range"] = [value_range[0], value_range[1]]
    return settings


def create_long_table(potentials, wavelengths, values):
    pot_grid, wl_grid = np.meshgrid(potentials, wavelengths)
    return pd.DataFrame({
        "Wavelength / nm": wl_grid.ravel(),
        "Potential / V": pot_grid.ravel(),
        "Value": values.ravel()
    })



def create_waterfall_tables(potentials, wavelengths, values, offset):
    """
    Creates Excel-ready waterfall tables.

    1) waterfall_shifted:
       Values exactly as plotted in the static waterfall diagram:
       y_shifted = y - min(y) + spectrum_index * offset

    2) waterfall_raw_processed:
       Processed but unshifted values.

    3) waterfall_offsets:
       Offset information for reproducing the waterfall plot in Excel.
    """
    shifted = {"Wavelength / nm": wavelengths}
    raw_processed = {"Wavelength / nm": wavelengths}
    offsets = []

    for j, pot in enumerate(potentials):
        y = values[:, j].astype(float)
        ymin = float(np.nanmin(y))
        vertical_offset = float(j * offset)
        col_name = f"{pot:g} V"

        raw_processed[col_name] = y
        shifted[col_name] = y - ymin + vertical_offset

        offsets.append({
            "Column": col_name,
            "Potential / V": pot,
            "Spectrum index": j,
            "Minimum subtracted": ymin,
            "Vertical offset added": vertical_offset,
            "Formula": "shifted = processed - minimum + vertical offset"
        })

    return (
        pd.DataFrame(shifted),
        pd.DataFrame(raw_processed),
        pd.DataFrame(offsets)
    )


# Backwards-compatible alias
def create_waterfall_table(potentials, wavelengths, values, offset):
    shifted, raw_processed, offsets = create_waterfall_tables(
        potentials, wavelengths, values, offset
    )
    return shifted


def create_waterfall_png(path, out_dir, mode_name, potentials, wavelengths, values,
                         wl_start, wl_final, offset, y_max=None):
    """
    Static waterfall plot for embedding into Excel.
    """
    png = out_dir / f"{path.stem}_{mode_name}_waterfall_excel_{safe_name_number(wl_start)}_to_{safe_name_number(wl_final)}.png"

    wf_df, _, _ = create_waterfall_tables(potentials, wavelengths, values, offset)
    cols = list(wf_df.columns[1:])

    plt.figure(figsize=(12, 8))
    colors = plt.cm.viridis(np.linspace(0, 1, max(len(cols), 1)))

    for i, col in enumerate(cols):
        plt.plot(wf_df["Wavelength / nm"], wf_df[col], lw=1.1, color=colors[i])

    plt.xlim(wl_start, wl_final)
    if y_max is not None:
        plt.ylim(0, y_max)

    ax = plt.gca()
    ax.xaxis.set_major_locator(MultipleLocator(50))
    ax.xaxis.set_minor_locator(MultipleLocator(10))
    ax.grid(True, which="major", axis="x", linestyle="--", alpha=0.35)

    plt.xlabel("Wavelength / nm")
    plt.ylabel("Value + vertical offset / a.u.")
    plt.title(f"{mode_name}: Waterfall plot")
    plt.tight_layout()
    plt.savefig(png, dpi=300)
    plt.close()

    return png, wf_df

def export_excel(path, out_dir, mode_name, potentials, wavelengths, raw, processed,
                 wl_start, wl_final, processing_text,
                 waterfall_shifted_df=None,
                 waterfall_unshifted_df=None,
                 waterfall_offsets_df=None,
                 waterfall_png=None):
    excel = out_dir / f"{path.stem}_{mode_name}_wavelength_{safe_name_number(wl_start)}_to_{safe_name_number(wl_final)}.xlsx"

    raw_matrix = pd.DataFrame(raw, index=wavelengths, columns=potentials)
    raw_matrix.index.name = "Wavelength / nm"
    raw_matrix.columns.name = "Potential / V"

    processed_matrix = pd.DataFrame(processed, index=wavelengths, columns=potentials)
    processed_matrix.index.name = "Wavelength / nm"
    processed_matrix.columns.name = "Potential / V"

    metadata = pd.DataFrame({
        "Property": [
            "Source file", "Mode", "Wavelength start / nm", "Wavelength final / nm",
            "Potential min / V", "Potential max / V",
            "Number of wavelengths", "Number of potentials",
            "Processing"
        ],
        "Value": [
            str(path), mode_name, wl_start, wl_final,
            float(np.nanmin(potentials)), float(np.nanmax(potentials)),
            len(wavelengths), len(potentials), processing_text
        ]
    })

    with pd.ExcelWriter(excel, engine="openpyxl") as writer:
        metadata.to_excel(writer, sheet_name="Metadata", index=False)
        raw_matrix.to_excel(writer, sheet_name="Raw_Matrix")
        processed_matrix.to_excel(writer, sheet_name="Processed_Matrix")
        create_long_table(potentials, wavelengths, raw).to_excel(writer, sheet_name="Raw_Long_Format", index=False)
        create_long_table(potentials, wavelengths, processed).to_excel(writer, sheet_name="Processed_Long_Format", index=False)
        if waterfall_shifted_df is not None:
            waterfall_shifted_df.to_excel(writer, sheet_name="Waterfall_Shifted_Values", index=False)
        if waterfall_unshifted_df is not None:
            waterfall_unshifted_df.to_excel(writer, sheet_name="Waterfall_Unshifted_Values", index=False)
        if waterfall_offsets_df is not None:
            waterfall_offsets_df.to_excel(writer, sheet_name="Waterfall_Offsets", index=False)

    if waterfall_png is not None and Path(waterfall_png).exists():
        wb = load_workbook(excel)
        if "Waterfall_Plot" in wb.sheetnames:
            del wb["Waterfall_Plot"]
        ws = wb.create_sheet("Waterfall_Plot", 0)
        ws["A1"] = f"{mode_name}: Waterfall plot"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A2"] = "Static waterfall plot generated from Waterfall_Shifted_Values."
        ws["A3"] = f"Wavelength range: {wl_start:g} to {wl_final:g} nm"
        img = ExcelImage(str(waterfall_png))
        img.width = 1100
        img.height = 750
        ws.add_image(img, "A5")
        wb.active = 0
        wb.save(excel)

    return excel


def plot_surface(go, path, out_dir, mode_name, potentials, wavelengths, z, wl_start, wl_final, suffix, label, z_range):
    fig = go.Figure()
    fig.add_trace(go.Surface(
        x=potentials, y=wavelengths, z=z,
        colorscale="Turbo",
        colorbar=dict(title=label),
        contours={"z": {"show": True, "usecolormap": True, "project_z": True}}
    ))
    fig.update_layout(
        title=f"{mode_name}: Interactive 3D voltammogram ({wl_start:g}-{wl_final:g} nm)",
        scene=dict(
            xaxis=dict(title="Potential / V"),
            yaxis=axis_settings_wavelength(wl_start, wl_final),
            zaxis=axis_settings_intensity(label, z_range),
            camera=dict(eye=dict(x=1.55, y=-1.75, z=1.10)),
            aspectmode="manual",
            aspectratio=dict(x=1.4, y=1.1, z=0.75)
        ),
        width=1150, height=820, margin=dict(l=0, r=0, b=0, t=55)
    )
    html = out_dir / f"{path.stem}_{mode_name}_surface_{safe_name_number(wl_start)}_to_{safe_name_number(wl_final)}_{suffix}.html"
    fig.write_html(html, include_plotlyjs="cdn")
    return fig, html


def plot_heatmap(go, path, out_dir, mode_name, potentials, wavelengths, z, wl_start, wl_final, suffix, label, z_range):
    kwargs = dict(x=potentials, y=wavelengths, z=z, colorscale="Turbo", colorbar=dict(title=label))
    if z_range is not None:
        kwargs["zmin"] = z_range[0]
        kwargs["zmax"] = z_range[1]
    fig = go.Figure(go.Heatmap(**kwargs))
    fig.update_layout(
        title=f"{mode_name}: Heatmap ({wl_start:g}-{wl_final:g} nm)",
        xaxis_title="Potential / V",
        yaxis=axis_settings_wavelength(wl_start, wl_final),
        width=1050, height=760, margin=dict(l=70, r=20, b=60, t=70)
    )
    html = out_dir / f"{path.stem}_{mode_name}_heatmap_{safe_name_number(wl_start)}_to_{safe_name_number(wl_final)}_{suffix}.html"
    fig.write_html(html, include_plotlyjs="cdn")
    return fig, html


def plot_contour(go, path, out_dir, mode_name, potentials, wavelengths, z, wl_start, wl_final, suffix, label, z_range):
    kwargs = dict(
        x=potentials, y=wavelengths, z=z, colorscale="Turbo",
        contours=dict(coloring="heatmap", showlabels=True, labelfont=dict(size=10)),
        colorbar=dict(title=label)
    )
    if z_range is not None:
        kwargs["zmin"] = z_range[0]
        kwargs["zmax"] = z_range[1]
    fig = go.Figure(go.Contour(**kwargs))
    fig.update_layout(
        title=f"{mode_name}: Contour plot ({wl_start:g}-{wl_final:g} nm)",
        xaxis_title="Potential / V",
        yaxis=axis_settings_wavelength(wl_start, wl_final),
        width=1050, height=760, margin=dict(l=70, r=20, b=60, t=70)
    )
    html = out_dir / f"{path.stem}_{mode_name}_contour_{safe_name_number(wl_start)}_to_{safe_name_number(wl_final)}_{suffix}.html"
    fig.write_html(html, include_plotlyjs="cdn")
    return fig, html


def plot_waterfall(go, path, out_dir, mode_name, potentials, wavelengths, z, wl_start, wl_final, suffix, label, z_range):
    fig = go.Figure()
    max_lines = 120
    indices = np.linspace(0, len(potentials) - 1, max_lines, dtype=int) if len(potentials) > max_lines else np.arange(len(potentials))

    for idx in indices:
        pot = potentials[idx]
        fig.add_trace(go.Scatter3d(
            x=wavelengths,
            y=np.full_like(wavelengths, pot, dtype=float),
            z=z[:, idx],
            mode="lines",
            line=dict(width=3),
            name=f"{pot:g} V",
            showlegend=False,
            hovertemplate="Wavelength: %{x:.2f} nm<br>Potential: %{y:.4g} V<br>Value: %{z:.4g}<extra></extra>"
        ))

    fig.update_layout(
        title=f"{mode_name}: Rotatable waterfall plot ({wl_start:g}-{wl_final:g} nm)",
        scene=dict(
            xaxis=axis_settings_wavelength(wl_start, wl_final),
            yaxis=dict(title="Potential / V"),
            zaxis=axis_settings_intensity(label, z_range),
            camera=dict(eye=dict(x=1.55, y=-1.9, z=1.1)),
            aspectmode="manual",
            aspectratio=dict(x=1.7, y=1.1, z=0.85)
        ),
        width=1150, height=820, margin=dict(l=0, r=0, b=0, t=55)
    )
    html = out_dir / f"{path.stem}_{mode_name}_waterfall_{safe_name_number(wl_start)}_to_{safe_name_number(wl_final)}_{suffix}.html"
    fig.write_html(html, include_plotlyjs="cdn")
    return fig, html


class SpectroVoltammogramGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Absorpto-/Fluorovoltammogramm Analysis V4")
        self.root.geometry("820x760")

        self.csv_path = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.mode = tk.StringVar(value="Absorptovoltammogramm")

        self.wl_start = tk.StringVar(value="300")
        self.wl_final = tk.StringVar(value="800")

        self.smoothing = tk.StringVar(value="savgol")
        self.savgol_window = tk.StringVar(value="11")
        self.savgol_poly = tk.StringVar(value="3")
        self.moving_window = tk.StringVar(value="5")

        self.scaling = tk.StringVar(value="raw")
        self.auto_axis = tk.BooleanVar(value=True)
        self.axis_min = tk.StringVar()
        self.axis_max = tk.StringVar()
        self.waterfall_offset = tk.StringVar(value="0.10")

        self.create_surface = tk.BooleanVar(value=True)
        self.create_heatmap = tk.BooleanVar(value=True)
        self.create_contour = tk.BooleanVar(value=True)
        self.create_waterfall = tk.BooleanVar(value=True)

        self.create_widgets()

    def create_widgets(self):
        pad = {"padx": 8, "pady": 5}

        frm = ttk.LabelFrame(self.root, text="Input and output")
        frm.pack(fill="x", **pad)

        ttk.Label(frm, text="CSV file:").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(frm, textvariable=self.csv_path, width=78).grid(row=0, column=1, **pad)
        ttk.Button(frm, text="Browse", command=self.browse_csv).grid(row=0, column=2, **pad)

        ttk.Label(frm, text="Output folder:").grid(row=1, column=0, sticky="w", **pad)
        ttk.Entry(frm, textvariable=self.output_dir, width=78).grid(row=1, column=1, **pad)
        ttk.Button(frm, text="Browse", command=self.browse_output).grid(row=1, column=2, **pad)

        ttk.Label(frm, text="Mode:").grid(row=2, column=0, sticky="w", **pad)
        ttk.Radiobutton(frm, text="Absorptovoltammogramm", variable=self.mode, value="Absorptovoltammogramm").grid(row=2, column=1, sticky="w", **pad)
        ttk.Radiobutton(frm, text="Fluorovoltammogramm", variable=self.mode, value="Fluorovoltammogramm").grid(row=2, column=2, sticky="w", **pad)

        frm_range = ttk.LabelFrame(self.root, text="Wavelength range")
        frm_range.pack(fill="x", **pad)
        ttk.Label(frm_range, text="Start / nm:").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(frm_range, textvariable=self.wl_start, width=12).grid(row=0, column=1, sticky="w", **pad)
        ttk.Label(frm_range, text="Final / nm:").grid(row=0, column=2, sticky="w", **pad)
        ttk.Entry(frm_range, textvariable=self.wl_final, width=12).grid(row=0, column=3, sticky="w", **pad)
        ttk.Button(frm_range, text="Read CSV range", command=self.read_range).grid(row=0, column=4, **pad)

        frm_proc = ttk.LabelFrame(self.root, text="Smoothing and scaling")
        frm_proc.pack(fill="x", **pad)

        ttk.Radiobutton(frm_proc, text="No smoothing", variable=self.smoothing, value="none").grid(row=0, column=0, sticky="w", **pad)
        ttk.Radiobutton(frm_proc, text="Moving average", variable=self.smoothing, value="moving").grid(row=1, column=0, sticky="w", **pad)
        ttk.Label(frm_proc, text="Window:").grid(row=1, column=1, sticky="e", **pad)
        ttk.Entry(frm_proc, textvariable=self.moving_window, width=8).grid(row=1, column=2, sticky="w", **pad)
        ttk.Radiobutton(frm_proc, text="Savitzky-Golay", variable=self.smoothing, value="savgol").grid(row=2, column=0, sticky="w", **pad)
        ttk.Label(frm_proc, text="Window:").grid(row=2, column=1, sticky="e", **pad)
        ttk.Entry(frm_proc, textvariable=self.savgol_window, width=8).grid(row=2, column=2, sticky="w", **pad)
        ttk.Label(frm_proc, text="Polynomial:").grid(row=2, column=3, sticky="e", **pad)
        ttk.Entry(frm_proc, textvariable=self.savgol_poly, width=8).grid(row=2, column=4, sticky="w", **pad)

        ttk.Radiobutton(frm_proc, text="Raw", variable=self.scaling, value="raw").grid(row=3, column=0, sticky="w", **pad)
        ttk.Radiobutton(frm_proc, text="Clip 1-99%", variable=self.scaling, value="clip99").grid(row=3, column=1, sticky="w", **pad)
        ttk.Radiobutton(frm_proc, text="log1p", variable=self.scaling, value="log1p").grid(row=3, column=2, sticky="w", **pad)
        ttk.Radiobutton(frm_proc, text="Clip + log1p", variable=self.scaling, value="clip99_log1p").grid(row=3, column=3, sticky="w", **pad)

        frm_axis = ttk.LabelFrame(self.root, text="Intensity / absorption axis")
        frm_axis.pack(fill="x", **pad)
        ttk.Checkbutton(frm_axis, text="Automatic axis", variable=self.auto_axis).grid(row=0, column=0, sticky="w", **pad)
        ttk.Label(frm_axis, text="Minimum:").grid(row=0, column=1, sticky="e", **pad)
        ttk.Entry(frm_axis, textvariable=self.axis_min, width=12).grid(row=0, column=2, sticky="w", **pad)
        ttk.Label(frm_axis, text="Maximum:").grid(row=0, column=3, sticky="e", **pad)
        ttk.Entry(frm_axis, textvariable=self.axis_max, width=12).grid(row=0, column=4, sticky="w", **pad)

        ttk.Label(frm_axis, text="Waterfall vertical offset:").grid(row=1, column=0, sticky="w", **pad)
        ttk.Entry(frm_axis, textvariable=self.waterfall_offset, width=12).grid(row=1, column=1, sticky="w", **pad)

        frm_plots = ttk.LabelFrame(self.root, text="Plots")
        frm_plots.pack(fill="x", **pad)
        ttk.Checkbutton(frm_plots, text="3D Surface", variable=self.create_surface).grid(row=0, column=0, sticky="w", **pad)
        ttk.Checkbutton(frm_plots, text="Heatmap", variable=self.create_heatmap).grid(row=0, column=1, sticky="w", **pad)
        ttk.Checkbutton(frm_plots, text="Contour", variable=self.create_contour).grid(row=0, column=2, sticky="w", **pad)
        ttk.Checkbutton(frm_plots, text="Waterfall", variable=self.create_waterfall).grid(row=0, column=3, sticky="w", **pad)

        ttk.Button(self.root, text="Create Excel and HTML files", command=self.run).pack(padx=8, pady=10)

        self.log = tk.Text(self.root, height=16, wrap="word")
        self.log.pack(fill="both", expand=True, padx=8, pady=8)

    def write_log(self, text):
        self.log.insert("end", text + "\n")
        self.log.see("end")
        self.root.update_idletasks()

    def browse_csv(self):
        filename = filedialog.askopenfilename(title="Select CSV file", filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if filename:
            self.csv_path.set(filename)
            self.output_dir.set(str(Path(filename).parent))
            name = Path(filename).name.lower()
            if "fluoro" in name:
                self.mode.set("Fluorovoltammogramm")
            elif "absor" in name:
                self.mode.set("Absorptovoltammogramm")
            self.read_range()

    def browse_output(self):
        dirname = filedialog.askdirectory(title="Select output folder")
        if dirname:
            self.output_dir.set(dirname)

    def read_range(self):
        try:
            p = Path(self.csv_path.get().strip().strip('"'))
            if not p.exists():
                messagebox.showerror("Error", "Please select a valid CSV file first.")
                return
            potentials, wavelengths, values, enc = read_voltammogram_csv(p)
            self.wl_start.set(f"{float(np.nanmin(wavelengths)):g}")
            self.wl_final.set(f"{float(np.nanmax(wavelengths)):g}")
            self.write_log(f"CSV loaded: {p.name}")
            self.write_log(f"Encoding: {enc}")
            self.write_log(f"Wavelength range: {self.wl_start.get()} to {self.wl_final.get()} nm")
            self.write_log(f"Potential range: {float(np.nanmin(potentials)):g} to {float(np.nanmax(potentials)):g} V")
            self.write_log(f"Matrix: {values.shape[0]} wavelengths x {values.shape[1]} potentials")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.write_log(f"ERROR: {e}")

    def process_values(self, raw):
        mode = self.smoothing.get()
        if mode == "none":
            return raw.copy(), "No smoothing"
        if mode == "moving":
            w = int(self.moving_window.get())
            return moving_average(raw, w), f"Moving average, {w} points"
        w = int(self.savgol_window.get())
        p = int(self.savgol_poly.get())
        return savgol_smooth(raw, w, p), f"Savitzky-Golay, window {w}, polynomial {p}"

    def axis_range(self):
        if self.auto_axis.get():
            return None
        lo = self.axis_min.get().strip().replace(",", ".")
        hi = self.axis_max.get().strip().replace(",", ".")
        if not lo or not hi:
            return None
        lo = float(lo)
        hi = float(hi)
        if lo > hi:
            lo, hi = hi, lo
        return (lo, hi)

    def run(self):
        try:
            path = Path(self.csv_path.get().strip().strip('"'))
            if not path.exists():
                messagebox.showerror("Error", "Please select a valid CSV file.")
                return

            out_dir = Path(self.output_dir.get().strip().strip('"')) if self.output_dir.get().strip() else path.parent
            out_dir.mkdir(parents=True, exist_ok=True)

            wl_start = float(self.wl_start.get().replace(",", "."))
            wl_final = float(self.wl_final.get().replace(",", "."))
            if wl_start > wl_final:
                wl_start, wl_final = wl_final, wl_start

            mode_name = self.mode.get()

            self.write_log("Reading CSV ...")
            potentials, wavelengths, raw, enc = read_voltammogram_csv(path)
            mask = (wavelengths >= wl_start) & (wavelengths <= wl_final)
            if not np.any(mask):
                raise ValueError("No data points in the selected wavelength range.")

            wavelengths = wavelengths[mask]
            raw = raw[mask, :]

            processed, processing_text = self.process_values(raw)
            z, label, suffix = scale_values(processed, self.scaling.get())
            z_range = self.axis_range()

            waterfall_offset = float(self.waterfall_offset.get().strip().replace(",", ".")) if self.waterfall_offset.get().strip() else 0.0

            self.write_log("Creating Excel-ready waterfall values ...")
            wf_shifted_df, wf_unshifted_df, wf_offsets_df = create_waterfall_tables(
                potentials, wavelengths, z, waterfall_offset
            )

            self.write_log("Creating static waterfall plot for Excel ...")
            wf_png, _ = create_waterfall_png(
                path, out_dir, mode_name, potentials, wavelengths, z,
                wl_start, wl_final, waterfall_offset,
                z_range[1] if z_range is not None else None
            )

            self.write_log("Writing Excel ...")
            excel = export_excel(
                path, out_dir, mode_name, potentials, wavelengths, raw, z,
                wl_start, wl_final, processing_text,
                waterfall_shifted_df=wf_shifted_df,
                waterfall_unshifted_df=wf_unshifted_df,
                waterfall_offsets_df=wf_offsets_df,
                waterfall_png=wf_png
            )

            self.write_log("Creating plots ...")
            go = ensure_plotly()
            outputs = [("Excel", excel), ("Waterfall PNG for Excel", wf_png)]

            if self.create_surface.get():
                outputs.append(("Surface", plot_surface(go, path, out_dir, mode_name, potentials, wavelengths, z, wl_start, wl_final, suffix, label, z_range)[1]))
            if self.create_heatmap.get():
                outputs.append(("Heatmap", plot_heatmap(go, path, out_dir, mode_name, potentials, wavelengths, z, wl_start, wl_final, suffix, label, z_range)[1]))
            if self.create_contour.get():
                outputs.append(("Contour", plot_contour(go, path, out_dir, mode_name, potentials, wavelengths, z, wl_start, wl_final, suffix, label, z_range)[1]))
            if self.create_waterfall.get():
                outputs.append(("Waterfall", plot_waterfall(go, path, out_dir, mode_name, potentials, wavelengths, z, wl_start, wl_final, suffix, label, z_range)[1]))

            for name, p in outputs:
                self.write_log(f"{name}: {p}")

            messagebox.showinfo("Finished", f"Files created in:\n{out_dir}")

        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.write_log(f"ERROR: {e}")


def main():
    root = tk.Tk()
    SpectroVoltammogramGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
