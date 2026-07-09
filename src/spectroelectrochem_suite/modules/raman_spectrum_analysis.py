"""
Raman-Auswertung GUI Version 14

Diese Version basiert auf "Raman_Auswertung_V13_Referenz_0_x" und wurde so umgebaut,
dass zu Beginn bequem eine CSV-Datei ausgewählt und der Wellenzahlbereich eingegeben wird.

Funktionen:
- CSV-Datei per Windows-Dialog auswählen
- Ausgabeordner auswählen
- Wellenzahlbereich eingeben
- Basislinienkorrektur mit modpoly
- Savitzky-Golay-Glättung
- Peakbestimmung mit FWHM
- optionale Peakidentifikation:
    0 = keine Referenz
    x = eigene Referenz-Excel-Datei auswählen
- Excel-Ausgabe mit:
    01_2D_Wasserfall_Grafik
    02_Korrigierte_Spektren
    03_2D_Wasserfall_Daten
    04_Peakparameter
    05_Peakidentifikation
    06_Verwendete_Referenz
- PNG- und PDF-Ausgabe des 2D-Wasserfalldiagramms
- Kontrolle der geglätteten/basiskorrigierten Spektren

CSV-Struktur:
- erste Spalte: Wellenzahlen / Raman shift
- weitere Spalten: Ramanintensitäten der einzelnen Spektren

Benötigte Pakete:
    py -m pip install pandas numpy matplotlib scipy pybaselines openpyxl
"""

from pathlib import Path
import sys
import subprocess
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.ticker import MultipleLocator

from scipy.signal import savgol_filter, find_peaks, peak_widths
from pybaselines import Baseline

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font


def ensure_plotly():
    try:
        import plotly.graph_objects as go
        return go
    except ModuleNotFoundError:
        answer = messagebox.askyesno(
            "Missing package",
            "The package 'plotly' is missing. It is required for the rotatable waterfall plot.\n\nInstall it now?"
        )
        if answer:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "plotly"])
            import plotly.graph_objects as go
            return go
        raise SystemExit("Missing package: plotly")


def ensure_package(import_name, pip_name=None):
    if pip_name is None:
        pip_name = import_name
    try:
        return __import__(import_name)
    except ModuleNotFoundError:
        answer = messagebox.askyesno(
            "Missing package",
            f"The package '{pip_name}' is missing.\n\nInstall it now?"
        )
        if answer:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])
            return __import__(import_name)
        raise SystemExit(f"Missing package: {pip_name}")


def safe_float(value):
    if pd.isna(value):
        return np.nan
    text = str(value).strip()
    if text == "":
        return np.nan
    text = text.replace(",", ".")
    return pd.to_numeric(text, errors="coerce")


def lese_csv(dateipfad: Path):
    """
    Liest eine Raman-CSV robust ein.
    Akzeptiert typische Windows-/Excel-Codierungen und unterschiedliche Trennzeichen.
    """
    encodings = ["utf-8-sig", "cp1252", "latin1", "utf-16", "utf-16le", "utf-16be"]
    last_error = None

    for enc in encodings:
        try:
            df_raw = pd.read_csv(
                dateipfad,
                sep=None,
                engine="python",
                dtype=str,
                encoding=enc
            )
            encoding_used = enc
            break
        except UnicodeDecodeError as err:
            last_error = err
    else:
        raise UnicodeDecodeError(
            "unknown", b"", 0, 1,
            "Die CSV-Datei konnte nicht mit üblichen Codierungen gelesen werden."
        ) from last_error

    # Numerische Konvertierung, Dezimalkomma wird akzeptiert
    df = df_raw.copy()
    for col in df.columns:
        df[col] = df[col].map(safe_float)

    # Zeilen ohne Wellenzahl entfernen
    df = df.dropna(subset=[df.columns[0]])
    df = df.dropna(axis=1, how="all")

    if df.shape[1] < 2:
        raise ValueError("Die CSV-Datei muss mindestens zwei Spalten enthalten: Wellenzahl + Intensität.")

    return df, encoding_used


def standardisiere_referenz_df(referenz_df):
    benoetigte_spalten = [
        "Peak_cm-1", "Intensitaet", "Schwingung", "Zuordnung",
        "Literatur", "DOI", "Kommentar"
    ]

    for spalte in benoetigte_spalten:
        if spalte not in referenz_df.columns:
            referenz_df[spalte] = ""

    referenz_df["Peak_cm-1"] = pd.to_numeric(referenz_df["Peak_cm-1"], errors="coerce")
    referenz_df = referenz_df.dropna(subset=["Peak_cm-1"])

    return referenz_df[benoetigte_spalten]


def korrigiere_spektrum(x_valid, y_valid, baseline_method, poly_order,
                       baseline_lambda, baseline_max_iter,
                       savgol_fenster, savgol_polynom):
    """Baseline correction with ModPoly, AsLS, arPLS or airPLS + Savitzky-Golay smoothing."""
    base = Baseline(x_data=x_valid)
    method = str(baseline_method).lower()

    if method == "modpoly":
        baseline, params = base.modpoly(y_valid, poly_order=int(poly_order), max_iter=int(baseline_max_iter))
        method_text = f"ModPoly, polynomial order {poly_order}"
    elif method == "asls":
        baseline, params = base.asls(y_valid, lam=float(baseline_lambda), p=0.01, max_iter=int(baseline_max_iter))
        method_text = f"AsLS, lambda={baseline_lambda:g}, p=0.01"
    elif method == "airpls":
        try:
            baseline, params = base.airpls(y_valid, lam=float(baseline_lambda), max_iter=int(baseline_max_iter))
        except Exception:
            baseline, params = base.asls(y_valid, lam=float(baseline_lambda), p=0.001, max_iter=int(baseline_max_iter))
        method_text = f"airPLS, lambda={baseline_lambda:g}"
    else:
        try:
            baseline, params = base.arpls(y_valid, lam=float(baseline_lambda), max_iter=int(baseline_max_iter))
        except Exception:
            baseline, params = base.asls(y_valid, lam=float(baseline_lambda), p=0.01, max_iter=int(baseline_max_iter))
        method_text = f"arPLS, lambda={baseline_lambda:g}"

    y_corr = y_valid - baseline

    n = len(y_corr)
    window = int(savgol_fenster)
    if window % 2 == 0:
        window += 1
    if window > n:
        window = n if n % 2 == 1 else n - 1
    if window <= int(savgol_polynom):
        window = int(savgol_polynom) + 2
        if window % 2 == 0:
            window += 1

    if window < 5 or window > n:
        y_smooth = y_corr.copy()
    else:
        y_smooth = savgol_filter(y_corr, window_length=window, polyorder=int(savgol_polynom), mode="interp")

    return baseline, y_corr, y_smooth, method_text


def bestimme_peaks(spektrum_name, x_valid, y_smooth, peak_prominence, peak_distance):
    peak_liste = []

    peaks, properties = find_peaks(
        y_smooth,
        prominence=peak_prominence,
        distance=peak_distance
    )

    if len(peaks) == 0:
        peak_liste.append({
            "Spektrum": spektrum_name,
            "Peak_Nr": np.nan,
            "Peaklage_cm-1": np.nan,
            "Intensitaet": np.nan,
            "FWHM_cm-1": np.nan
        })
        return peak_liste

    widths_result = peak_widths(
        y_smooth,
        peaks,
        rel_height=0.5
    )

    widths_points = widths_result[0]
    dx = np.mean(np.abs(np.diff(x_valid)))
    widths_cm = widths_points * dx

    for nummer, peak_index in enumerate(peaks, start=1):
        peak_liste.append({
            "Spektrum": spektrum_name,
            "Peak_Nr": nummer,
            "Peaklage_cm-1": x_valid[peak_index],
            "Intensitaet": y_smooth[peak_index],
            "FWHM_cm-1": widths_cm[nummer - 1]
        })

    return peak_liste


def identifiziere_peaks(peaks_df, referenz_df, toleranz, substanzname):
    identifikationen = []

    for _, peak in peaks_df.iterrows():
        gemessen = peak["Peaklage_cm-1"]

        if pd.isna(gemessen):
            identifikationen.append({
                "Substanz": substanzname,
                "Spektrum": peak["Spektrum"],
                "Peak_Nr": peak["Peak_Nr"],
                "Gemessen_cm-1": np.nan,
                "Intensitaet": peak["Intensitaet"],
                "FWHM_cm-1": peak["FWHM_cm-1"],
                "Referenz_cm-1": np.nan,
                "Abweichung_cm-1": np.nan,
                "Betrag_Abweichung_cm-1": np.nan,
                "Bewertung": "kein Peak gefunden",
                "Referenz_Intensitaet": "",
                "Schwingung": "",
                "Zuordnung": "",
                "Literatur": "",
                "DOI": "",
                "Kommentar": ""
            })
            continue

        treffer = []

        for _, ref in referenz_df.iterrows():
            ref_peak = ref["Peak_cm-1"]
            abweichung = gemessen - ref_peak

            if abs(abweichung) <= toleranz:
                betrag = abs(abweichung)

                if betrag <= toleranz / 4:
                    bewertung = "sehr gute Übereinstimmung"
                elif betrag <= toleranz / 2:
                    bewertung = "gute Übereinstimmung"
                else:
                    bewertung = "mögliche Übereinstimmung"

                treffer.append({
                    "Substanz": substanzname,
                    "Spektrum": peak["Spektrum"],
                    "Peak_Nr": peak["Peak_Nr"],
                    "Gemessen_cm-1": gemessen,
                    "Intensitaet": peak["Intensitaet"],
                    "FWHM_cm-1": peak["FWHM_cm-1"],
                    "Referenz_cm-1": ref_peak,
                    "Abweichung_cm-1": abweichung,
                    "Betrag_Abweichung_cm-1": betrag,
                    "Bewertung": bewertung,
                    "Referenz_Intensitaet": ref.get("Intensitaet", ""),
                    "Schwingung": ref.get("Schwingung", ""),
                    "Zuordnung": ref.get("Zuordnung", ""),
                    "Literatur": ref.get("Literatur", ""),
                    "DOI": ref.get("DOI", ""),
                    "Kommentar": ref.get("Kommentar", "")
                })

        if len(treffer) == 0:
            identifikationen.append({
                "Substanz": substanzname,
                "Spektrum": peak["Spektrum"],
                "Peak_Nr": peak["Peak_Nr"],
                "Gemessen_cm-1": gemessen,
                "Intensitaet": peak["Intensitaet"],
                "FWHM_cm-1": peak["FWHM_cm-1"],
                "Referenz_cm-1": np.nan,
                "Abweichung_cm-1": np.nan,
                "Betrag_Abweichung_cm-1": np.nan,
                "Bewertung": "nicht zugeordnet",
                "Referenz_Intensitaet": "",
                "Schwingung": "",
                "Zuordnung": "",
                "Literatur": "",
                "DOI": "",
                "Kommentar": f"kein Referenzpeak innerhalb ±{toleranz} cm-1"
            })
        else:
            treffer = sorted(treffer, key=lambda x: x["Betrag_Abweichung_cm-1"])
            identifikationen.extend(treffer)

    return pd.DataFrame(identifikationen)


def leere_identifikationstabelle(peaks_df, substanzname):
    zeilen = []

    for _, peak in peaks_df.iterrows():
        zeilen.append({
            "Substanz": substanzname,
            "Spektrum": peak["Spektrum"],
            "Peak_Nr": peak["Peak_Nr"],
            "Gemessen_cm-1": peak["Peaklage_cm-1"],
            "Intensitaet": peak["Intensitaet"],
            "FWHM_cm-1": peak["FWHM_cm-1"],
            "Referenz_cm-1": np.nan,
            "Abweichung_cm-1": np.nan,
            "Betrag_Abweichung_cm-1": np.nan,
            "Bewertung": "keine Referenz verwendet",
            "Referenz_Intensitaet": "",
            "Schwingung": "",
            "Zuordnung": "",
            "Literatur": "",
            "DOI": "",
            "Kommentar": "Keine Referenz ausgewählt"
        })

    return pd.DataFrame(zeilen)


def erstelle_2d_wasserfall_plot(wasserfall_df, substanzname, bereich_min, bereich_max,
                                ausgabe_png, ausgabe_pdf, max_intensity=None):
    plt.figure(figsize=(12, 8))

    x_plot = wasserfall_df["Raman_shift_cm-1"].to_numpy()
    spektrum_spalten = list(wasserfall_df.columns[1:])
    anzahl_spektren = len(spektrum_spalten)

    farben = plt.cm.coolwarm(np.linspace(0, 1, max(anzahl_spektren, 1)))

    for i, col in enumerate(spektrum_spalten):
        y_plot = wasserfall_df[col].to_numpy()
        plt.plot(x_plot, y_plot, lw=1, color=farben[i])

    plt.xlim(bereich_min, bereich_max)

    if max_intensity is not None:
        plt.ylim(0, max_intensity)

    ax = plt.gca()
    ax.xaxis.set_major_locator(MultipleLocator(100))
    ax.xaxis.set_minor_locator(MultipleLocator(20))

    ax.tick_params(axis="x", which="major", length=8, width=1.2, direction="inout")
    ax.tick_params(axis="x", which="minor", length=4, width=0.8, direction="inout")
    ax.grid(True, which="major", axis="x", linestyle="--", alpha=0.35)

    plt.xlabel("Raman shift / cm$^{-1}$")
    plt.ylabel("Corrected intensity + vertical offset / counts")
    plt.title(f"2D waterfall plot: {substanzname}")
    plt.tight_layout()

    plt.savefig(ausgabe_png, dpi=300)
    plt.savefig(ausgabe_pdf)
    plt.close()



def erstelle_drehbares_wasserfall_plot(wasserfall_df, substanzname, bereich_min, bereich_max,
                                      ausgabe_html, max_intensity=None):
    """
    Erstellt ein interaktives, drehbares Wasserfalldiagramm als HTML-Datei.
    X = Raman shift, Y = spectrum number, Z = corrected intensity + offset.
    """
    go = ensure_plotly()

    x = wasserfall_df["Raman_shift_cm-1"].to_numpy()
    spektrum_spalten = list(wasserfall_df.columns[1:])

    fig = go.Figure()

    for i, col in enumerate(spektrum_spalten, start=1):
        y_vals = np.full_like(x, i, dtype=float)
        z_vals = wasserfall_df[col].to_numpy(dtype=float)

        fig.add_trace(
            go.Scatter3d(
                x=x,
                y=y_vals,
                z=z_vals,
                mode="lines",
                line=dict(width=3),
                name=str(col),
                showlegend=False,
                hovertemplate=(
                    "Raman shift: %{x:.2f} cm^-1<br>"
                    "Spectrum number: %{y}<br>"
                    "Intensity: %{z:.4g}<extra></extra>"
                )
            )
        )

    zaxis = dict(title="Corrected intensity + vertical offset / counts")
    if max_intensity is not None:
        zaxis["range"] = [0, max_intensity]

    fig.update_layout(
        title=f"Rotatable 3D waterfall plot: {substanzname}",
        scene=dict(
            xaxis=dict(
                title="Raman shift / cm^-1",
                tickmode="linear",
                tick0=100 * np.floor(min(bereich_min, bereich_max) / 100),
                dtick=100
            ),
            yaxis=dict(title="Spectrum number"),
            zaxis=zaxis,
            camera=dict(eye=dict(x=1.6, y=-1.9, z=1.1)),
            aspectmode="manual",
            aspectratio=dict(x=1.7, y=1.1, z=0.85)
        ),
        width=1150,
        height=820,
        margin=dict(l=0, r=0, b=0, t=60)
    )

    fig.write_html(ausgabe_html, include_plotlyjs="cdn")
    return ausgabe_html


def plot_kontrolle(spektren_df, bereich_min, bereich_max, ausgabe_png):
    plt.figure(figsize=(10, 5))

    for col in spektren_df.columns[1:]:
        plt.plot(
            spektren_df["Raman_shift_cm-1"],
            spektren_df[col],
            lw=1
        )

    plt.xlim(bereich_min, bereich_max)

    ax = plt.gca()
    ax.xaxis.set_major_locator(MultipleLocator(100))
    ax.xaxis.set_minor_locator(MultipleLocator(20))
    ax.tick_params(axis="x", which="major", length=8, width=1.2, direction="inout")
    ax.tick_params(axis="x", which="minor", length=4, width=0.8, direction="inout")

    plt.xlabel("Raman shift / cm$^{-1}$")
    plt.ylabel("Corrected intensity / counts")
    plt.title("Baseline-corrected and smoothed Raman spectra")
    plt.tight_layout()

    plt.savefig(ausgabe_png, dpi=300)
    plt.close()



def baseline_quality_metrics(raw_df, baseline_df, corrected_df):
    rows = []
    for col in raw_df.columns[1:]:
        corr = corrected_df[col].to_numpy(dtype=float)
        finite = np.isfinite(corr)
        if not np.any(finite):
            rows.append({"Spektrum": col, "Median_corrected": np.nan, "Std_corrected": np.nan,
                         "Negative_fraction": np.nan, "Quality_note": "not evaluable"})
            continue
        c = corr[finite]
        med = float(np.nanmedian(c))
        std = float(np.nanstd(c))
        neg = float(np.mean(c < 0))
        if neg > 0.35:
            note = "possible overcorrection"
        elif abs(med) > 0.2 * max(std, 1e-12):
            note = "possible residual baseline"
        else:
            note = "acceptable"
        rows.append({"Spektrum": col, "Median_corrected": med, "Std_corrected": std,
                     "Negative_fraction": neg, "Quality_note": note})
    return pd.DataFrame(rows)


def erstelle_baseline_kontrollplot(x, raw_df, baseline_df, corrected_df, smoothed_df,
                                  spektrum_name, bereich_min, bereich_max, ausgabe_png):
    if spektrum_name not in raw_df.columns:
        spektrum_name = raw_df.columns[1]
    plt.figure(figsize=(12, 7))
    plt.plot(x, raw_df[spektrum_name], color="0.55", lw=1.0, label="Raw spectrum")
    plt.plot(x, baseline_df[spektrum_name], color="red", lw=1.5, label="Calculated baseline")
    plt.plot(x, corrected_df[spektrum_name], color="tab:blue", lw=1.2, label="Baseline corrected")
    plt.plot(x, smoothed_df[spektrum_name], color="tab:green", lw=1.5, label="Corrected + smoothed")
    plt.xlim(bereich_min, bereich_max)
    ax = plt.gca()
    ax.xaxis.set_major_locator(MultipleLocator(100))
    ax.xaxis.set_minor_locator(MultipleLocator(20))
    ax.grid(True, which="major", axis="x", linestyle="--", alpha=0.35)
    plt.xlabel("Raman shift / cm$^{-1}$")
    plt.ylabel("Intensity / counts")
    plt.title(f"Baseline correction control: {spektrum_name}")
    plt.legend()
    plt.tight_layout()
    plt.savefig(ausgabe_png, dpi=300)
    plt.close()

def speichere_excel_mit_bild(ausgabe_excel, ausgabe_wasserfall_png,
                             spektren_df, wasserfall_df, peaks_df, ident_df,
                             referenz_df, substanzname, referenzdatei,
                             referenz_aktiv, identifikations_toleranz,
                             raw_df=None, baseline_df=None, corrected_df=None,
                             quality_df=None, baseline_control_png=None,
                             baseline_method_text=''):
    with pd.ExcelWriter(ausgabe_excel, engine="openpyxl") as writer:
        spektren_df.to_excel(writer, sheet_name="02_Korrigierte_Spektren", index=False)
        wasserfall_df.to_excel(writer, sheet_name="03_2D_Wasserfall_Daten", index=False)
        peaks_df.to_excel(writer, sheet_name="04_Peakparameter", index=False)
        ident_df.to_excel(writer, sheet_name="05_Peakidentifikation", index=False)

        if referenz_aktiv:
            referenz_df.to_excel(writer, sheet_name="06_Verwendete_Referenz", index=False)
        else:
            pd.DataFrame({"Hinweis": ["Keine Referenz verwendet."]}).to_excel(
                writer,
                sheet_name="06_Verwendete_Referenz",
                index=False
            )

        if raw_df is not None:
            raw_df.to_excel(writer, sheet_name="07_Raw_Spectra", index=False)
        if baseline_df is not None:
            baseline_df.to_excel(writer, sheet_name="08_Baselines", index=False)
        if corrected_df is not None:
            corrected_df.to_excel(writer, sheet_name="09_Baseline_Corrected", index=False)
        if quality_df is not None:
            quality_df.to_excel(writer, sheet_name="10_Baseline_Quality", index=False)

    wb = load_workbook(ausgabe_excel)

    if "01_2D_Wasserfall_Grafik" in wb.sheetnames:
        del wb["01_2D_Wasserfall_Grafik"]

    ws = wb.create_sheet("01_2D_Wasserfall_Grafik", 0)

    ws["A1"] = f"2D waterfall plot: {substanzname}"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "Python-generated PNG plot."
    ws["A3"] = f"Baseline method: {baseline_method_text}"

    if referenz_aktiv:
        ws["A3"] = f"Reference file: {referenzdatei}"
        ws["A4"] = f"Identification tolerance: ±{identifikations_toleranz} cm-1"
        ws["A5"] = "Peak identification: see sheet '05_Peakidentifikation'."
    else:
        ws["A3"] = "No reference file used."
        ws["A4"] = "Peak list without reference assignment: see sheet '04_Peakparameter'."

    img = ExcelImage(str(ausgabe_wasserfall_png))
    img.width = 1100
    img.height = 750
    ws.add_image(img, "A8")

    if baseline_control_png is not None and Path(baseline_control_png).exists():
        if "11_Baseline_Control_Plot" in wb.sheetnames:
            del wb["11_Baseline_Control_Plot"]
        ws2 = wb.create_sheet("11_Baseline_Control_Plot")
        ws2["A1"] = "Baseline correction control plot"
        ws2["A1"].font = Font(bold=True, size=16)
        ws2["A2"] = f"Baseline method: {baseline_method_text}"
        img2 = ExcelImage(str(baseline_control_png))
        img2.width = 1100
        img2.height = 700
        ws2.add_image(img2, "A4")

    for sheet in wb.worksheets:
        sheet.column_dimensions["A"].width = 28

    wb.active = 0
    wb.save(ausgabe_excel)


class RamanAuswertungGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Raman-Auswertung GUI V16 Scientific Edition")
        self.root.geometry("820x760")

        self.csv_path = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.substanzname = tk.StringVar(value="Unbekannte Substanz")

        self.wn_start = tk.StringVar(value="300")
        self.wn_final = tk.StringVar(value="1800")

        self.baseline_method = tk.StringVar(value="arpls")
        self.poly_order = tk.StringVar(value="5")
        self.baseline_lambda = tk.StringVar(value="100000")
        self.baseline_max_iter = tk.StringVar(value="100")
        self.savgol_window = tk.StringVar(value="21")
        self.savgol_poly = tk.StringVar(value="3")

        self.waterfall_offset = tk.StringVar(value="120")
        self.max_raman_intensity = tk.StringVar(value="")
        self.peak_prominence = tk.StringVar(value="20")
        self.peak_distance = tk.StringVar(value="10")

        self.reference_mode = tk.StringVar(value="none")
        self.reference_file = tk.StringVar()
        self.ident_tolerance = tk.StringVar(value="8.0")

        self.create_widgets()

    def create_widgets(self):
        pad = {"padx": 8, "pady": 5}

        frm_file = ttk.LabelFrame(self.root, text="Input and output")
        frm_file.pack(fill="x", **pad)

        ttk.Label(frm_file, text="CSV file:").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(frm_file, textvariable=self.csv_path, width=78).grid(row=0, column=1, **pad)
        ttk.Button(frm_file, text="Browse", command=self.browse_csv).grid(row=0, column=2, **pad)

        ttk.Label(frm_file, text="Output folder:").grid(row=1, column=0, sticky="w", **pad)
        ttk.Entry(frm_file, textvariable=self.output_dir, width=78).grid(row=1, column=1, **pad)
        ttk.Button(frm_file, text="Browse", command=self.browse_output).grid(row=1, column=2, **pad)

        ttk.Label(frm_file, text="Substance/sample name:").grid(row=2, column=0, sticky="w", **pad)
        ttk.Entry(frm_file, textvariable=self.substanzname, width=40).grid(row=2, column=1, sticky="w", **pad)

        frm_range = ttk.LabelFrame(self.root, text="Wavenumber range")
        frm_range.pack(fill="x", **pad)

        ttk.Label(frm_range, text="Start / cm^-1:").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(frm_range, textvariable=self.wn_start, width=12).grid(row=0, column=1, sticky="w", **pad)
        ttk.Label(frm_range, text="Final / cm^-1:").grid(row=0, column=2, sticky="w", **pad)
        ttk.Entry(frm_range, textvariable=self.wn_final, width=12).grid(row=0, column=3, sticky="w", **pad)
        ttk.Button(frm_range, text="Read CSV range", command=self.read_range_from_csv).grid(row=0, column=4, **pad)

        frm_corr = ttk.LabelFrame(self.root, text="Baseline correction and smoothing")
        frm_corr.pack(fill="x", **pad)

        ttk.Label(frm_corr, text="Baseline method:").grid(row=0, column=0, sticky="w", **pad)
        ttk.Radiobutton(frm_corr, text="ModPoly", variable=self.baseline_method, value="modpoly").grid(row=0, column=1, sticky="w", **pad)
        ttk.Radiobutton(frm_corr, text="AsLS", variable=self.baseline_method, value="asls").grid(row=0, column=2, sticky="w", **pad)
        ttk.Radiobutton(frm_corr, text="arPLS recommended", variable=self.baseline_method, value="arpls").grid(row=0, column=3, sticky="w", **pad)
        ttk.Radiobutton(frm_corr, text="airPLS fluorescence", variable=self.baseline_method, value="airpls").grid(row=0, column=4, sticky="w", **pad)

        ttk.Label(frm_corr, text="ModPoly order:").grid(row=1, column=0, sticky="w", **pad)
        ttk.Entry(frm_corr, textvariable=self.poly_order, width=10).grid(row=1, column=1, sticky="w", **pad)

        ttk.Label(frm_corr, text="Lambda:").grid(row=2, column=0, sticky="w", **pad)
        ttk.Entry(frm_corr, textvariable=self.baseline_lambda, width=12).grid(row=2, column=1, sticky="w", **pad)
        ttk.Label(frm_corr, text="Max iterations:").grid(row=2, column=2, sticky="w", **pad)
        ttk.Entry(frm_corr, textvariable=self.baseline_max_iter, width=10).grid(row=2, column=3, sticky="w", **pad)

        ttk.Label(frm_corr, text="Savitzky-Golay window:").grid(row=3, column=0, sticky="w", **pad)
        ttk.Entry(frm_corr, textvariable=self.savgol_window, width=10).grid(row=3, column=1, sticky="w", **pad)
        ttk.Label(frm_corr, text="Savitzky-Golay polynomial:").grid(row=3, column=2, sticky="w", **pad)
        ttk.Entry(frm_corr, textvariable=self.savgol_poly, width=10).grid(row=3, column=3, sticky="w", **pad)

        frm_peak = ttk.LabelFrame(self.root, text="Waterfall and peak detection")
        frm_peak.pack(fill="x", **pad)

        ttk.Label(frm_peak, text="Waterfall offset:").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(frm_peak, textvariable=self.waterfall_offset, width=10).grid(row=0, column=1, sticky="w", **pad)

        ttk.Label(frm_peak, text="Maximum Raman intensity:").grid(row=0, column=2, sticky="w", **pad)
        ttk.Entry(frm_peak, textvariable=self.max_raman_intensity, width=12).grid(row=0, column=3, sticky="w", **pad)
        ttk.Label(frm_peak, text="empty = automatic").grid(row=0, column=4, sticky="w", **pad)

        ttk.Label(frm_peak, text="Peak prominence:").grid(row=1, column=0, sticky="w", **pad)
        ttk.Entry(frm_peak, textvariable=self.peak_prominence, width=10).grid(row=1, column=1, sticky="w", **pad)

        ttk.Label(frm_peak, text="Peak distance / points:").grid(row=1, column=2, sticky="w", **pad)
        ttk.Entry(frm_peak, textvariable=self.peak_distance, width=10).grid(row=1, column=3, sticky="w", **pad)

        frm_ref = ttk.LabelFrame(self.root, text="Reference / peak assignment")
        frm_ref.pack(fill="x", **pad)

        ttk.Radiobutton(frm_ref, text="No reference", variable=self.reference_mode, value="none").grid(row=0, column=0, sticky="w", **pad)
        ttk.Radiobutton(frm_ref, text="Use own Excel reference file", variable=self.reference_mode, value="custom").grid(row=0, column=1, sticky="w", **pad)

        ttk.Label(frm_ref, text="Reference file:").grid(row=1, column=0, sticky="w", **pad)
        ttk.Entry(frm_ref, textvariable=self.reference_file, width=62).grid(row=1, column=1, **pad)
        ttk.Button(frm_ref, text="Browse", command=self.browse_reference).grid(row=1, column=2, **pad)

        ttk.Label(frm_ref, text="Tolerance / cm^-1:").grid(row=2, column=0, sticky="w", **pad)
        ttk.Entry(frm_ref, textvariable=self.ident_tolerance, width=10).grid(row=2, column=1, sticky="w", **pad)

        frm_run = ttk.Frame(self.root)
        frm_run.pack(fill="x", **pad)
        ttk.Button(frm_run, text="Start analysis", command=self.run_analysis).pack(side="left", padx=8, pady=10)

        self.log = tk.Text(self.root, height=16, wrap="word")
        self.log.pack(fill="both", expand=True, padx=8, pady=8)

    def write_log(self, text):
        self.log.insert("end", text + "\n")
        self.log.see("end")
        self.root.update_idletasks()

    def browse_csv(self):
        filename = filedialog.askopenfilename(
            title="Select Raman CSV file",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.csv_path.set(filename)
            p = Path(filename)
            self.output_dir.set(str(p.parent))
            self.read_range_from_csv()

    def browse_output(self):
        dirname = filedialog.askdirectory(title="Select output folder")
        if dirname:
            self.output_dir.set(dirname)

    def browse_reference(self):
        filename = filedialog.askopenfilename(
            title="Select reference Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.reference_file.set(filename)
            self.reference_mode.set("custom")

    def read_range_from_csv(self):
        try:
            csv = Path(self.csv_path.get().strip().strip('"'))
            if not csv.exists():
                messagebox.showerror("Error", "Please select a valid CSV file first.")
                return

            data, encoding = lese_csv(csv)
            x = data.iloc[:, 0].to_numpy(dtype=float)

            self.wn_start.set(f"{float(np.nanmin(x)):g}")
            self.wn_final.set(f"{float(np.nanmax(x)):g}")

            self.write_log(f"CSV loaded: {csv.name}")
            self.write_log(f"Encoding used: {encoding}")
            self.write_log(f"Available wavenumber range: {self.wn_start.get()} to {self.wn_final.get()} cm^-1")
            self.write_log(f"Number of spectra: {data.shape[1] - 1}")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.write_log(f"ERROR: {e}")

    def load_reference(self):
        mode = self.reference_mode.get()

        if mode == "none":
            return "Keine Referenz", None, pd.DataFrame(), False

        ref_path = Path(self.reference_file.get().strip().strip('"'))

        if not ref_path.exists():
            raise FileNotFoundError("Reference mode is active, but no valid reference Excel file was selected.")

        ref_df = pd.read_excel(ref_path)
        ref_df = standardisiere_referenz_df(ref_df)

        substanz = ref_path.stem
        return substanz, ref_path, ref_df, True

    def run_analysis(self):
        try:
            csv = Path(self.csv_path.get().strip().strip('"'))
            if not csv.exists():
                messagebox.showerror("Error", "Please select a valid CSV file.")
                return

            out_dir = Path(self.output_dir.get().strip().strip('"')) if self.output_dir.get().strip() else csv.parent
            out_dir.mkdir(parents=True, exist_ok=True)

            substanzname = self.substanzname.get().strip() or csv.stem

            bereich_min = float(self.wn_start.get().replace(",", "."))
            bereich_max = float(self.wn_final.get().replace(",", "."))
            if bereich_min > bereich_max:
                bereich_min, bereich_max = bereich_max, bereich_min

            baseline_method = self.baseline_method.get()
            poly_order = int(self.poly_order.get())
            baseline_lambda = float(self.baseline_lambda.get().replace(",", "."))
            baseline_max_iter = int(float(self.baseline_max_iter.get().replace(",", ".")))
            savgol_fenster = int(self.savgol_window.get())
            savgol_polynom = int(self.savgol_poly.get())

            wasserfall_abstand = float(self.waterfall_offset.get().replace(",", "."))

            max_intensity = None
            max_intensity_text = self.max_raman_intensity.get().strip().replace(",", ".")
            if max_intensity_text:
                max_intensity = float(max_intensity_text)
                self.write_log(f"Maximum Raman intensity for waterfall plots: {max_intensity:g}")
            else:
                self.write_log("Maximum Raman intensity: automatic")

            peak_prominence = float(self.peak_prominence.get().replace(",", "."))
            peak_distance = int(float(self.peak_distance.get().replace(",", ".")))
            identifikations_toleranz = float(self.ident_tolerance.get().replace(",", "."))

            ausgabe_excel = out_dir / f"{csv.stem}_Raman_Auswertung_V16.xlsx"
            ausgabe_wasserfall_png = out_dir / f"{csv.stem}_2D_Wasserfall_V16.png"
            ausgabe_wasserfall_pdf = out_dir / f"{csv.stem}_2D_Wasserfall_V16.pdf"
            ausgabe_wasserfall_html = out_dir / f"{csv.stem}_3D_drehbarer_Wasserfall_V16.html"
            ausgabe_kontrolle_png = out_dir / f"{csv.stem}_Kontrollplot_korrigiert_geglaettet_V16.png"
            ausgabe_baseline_kontrolle_png = out_dir / f"{csv.stem}_Baseline_Kontrollplot_V16.png"

            self.write_log("Loading reference ...")
            ref_substanz, referenzdatei, referenz_df, referenz_aktiv = self.load_reference()

            if referenz_aktiv and (not substanzname or substanzname == "Unbekannte Substanz"):
                substanzname = ref_substanz

            self.write_log("Reading CSV ...")
            data, encoding = lese_csv(csv)
            self.write_log(f"Encoding used: {encoding}")

            x = data.iloc[:, 0].to_numpy(dtype=float)
            mask = (x >= bereich_min) & (x <= bereich_max)
            x_cut = x[mask]

            if len(x_cut) < 5:
                raise ValueError("The selected wavenumber range contains too few data points.")

            result_dict = {"Raman_shift_cm-1": x_cut}
            wasserfall_dict = {"Raman_shift_cm-1": x_cut}
            raw_dict = {"Raman_shift_cm-1": x_cut}
            baseline_dict = {"Raman_shift_cm-1": x_cut}
            corrected_dict = {"Raman_shift_cm-1": x_cut}

            alle_peaks = []
            gueltige_spektren = 0
            baseline_method_text_used = ""

            for spektrum_nr, col in enumerate(data.columns[1:], start=1):
                y = data[col].to_numpy(dtype=float)
                y_cut = y[mask]

                valid = np.isfinite(x_cut) & np.isfinite(y_cut)

                x_valid = x_cut[valid]
                y_valid = y_cut[valid]

                if len(y_valid) < max(5, savgol_fenster):
                    self.write_log(f"Skipped, too few valid values: {col}")
                    continue

                if np.nanmax(y_valid) == np.nanmin(y_valid):
                    self.write_log(f"Skipped, constant spectrum: {col}")
                    continue

                self.write_log(f"Processing: {col}")

                baseline, y_corr, y_smooth, baseline_method_text = korrigiere_spektrum(
                    x_valid, y_valid, baseline_method, poly_order, baseline_lambda,
                    baseline_max_iter, savgol_fenster, savgol_polynom
                )
                baseline_method_text_used = baseline_method_text

                raw_full = np.full_like(x_cut, np.nan, dtype=float)
                baseline_full = np.full_like(x_cut, np.nan, dtype=float)
                corrected_full = np.full_like(x_cut, np.nan, dtype=float)
                smooth_full = np.full_like(x_cut, np.nan, dtype=float)

                raw_full[valid] = y_valid
                baseline_full[valid] = baseline
                corrected_full[valid] = y_corr
                smooth_full[valid] = y_smooth

                raw_dict[col] = raw_full
                baseline_dict[col] = baseline_full
                corrected_dict[col] = corrected_full
                result_dict[col] = smooth_full

                y_wasserfall = smooth_full - np.nanmin(smooth_full) + gueltige_spektren * wasserfall_abstand
                wasserfall_dict[col] = y_wasserfall
                gueltige_spektren += 1

                peakdaten = bestimme_peaks(
                    col,
                    x_valid,
                    y_smooth,
                    peak_prominence,
                    peak_distance
                )
                alle_peaks.extend(peakdaten)

            if gueltige_spektren == 0:
                raise ValueError("No valid spectra were processed.")

            spektren_df = pd.DataFrame(result_dict)
            wasserfall_df = pd.DataFrame(wasserfall_dict)
            raw_df = pd.DataFrame(raw_dict)
            baseline_df = pd.DataFrame(baseline_dict)
            corrected_df = pd.DataFrame(corrected_dict)
            peaks_df = pd.DataFrame(alle_peaks)
            quality_df = baseline_quality_metrics(raw_df, baseline_df, corrected_df)

            if referenz_aktiv:
                ident_df = identifiziere_peaks(
                    peaks_df,
                    referenz_df,
                    identifikations_toleranz,
                    substanzname
                )
            else:
                ident_df = leere_identifikationstabelle(peaks_df, substanzname)

            self.write_log("Creating waterfall plot ...")
            erstelle_2d_wasserfall_plot(
                wasserfall_df,
                substanzname,
                bereich_min,
                bereich_max,
                ausgabe_wasserfall_png,
                ausgabe_wasserfall_pdf,
                max_intensity
            )

            self.write_log("Creating rotatable 3D waterfall plot ...")
            erstelle_drehbares_wasserfall_plot(
                wasserfall_df,
                substanzname,
                bereich_min,
                bereich_max,
                ausgabe_wasserfall_html,
                max_intensity
            )

            self.write_log("Creating control plot ...")
            plot_kontrolle(spektren_df, bereich_min, bereich_max, ausgabe_kontrolle_png)

            self.write_log("Creating baseline control plot ...")
            erstelle_baseline_kontrollplot(
                x_cut, raw_df, baseline_df, corrected_df, spektren_df,
                spektren_df.columns[1], bereich_min, bereich_max,
                ausgabe_baseline_kontrolle_png
            )

            self.write_log("Writing Excel file ...")
            speichere_excel_mit_bild(
                ausgabe_excel,
                ausgabe_wasserfall_png,
                spektren_df,
                wasserfall_df,
                peaks_df,
                ident_df,
                referenz_df,
                substanzname,
                referenzdatei,
                referenz_aktiv,
                identifikations_toleranz,
                raw_df,
                baseline_df,
                corrected_df,
                quality_df,
                ausgabe_baseline_kontrolle_png,
                baseline_method_text_used
            )

            self.write_log("Finished.")
            self.write_log(f"Excel file: {ausgabe_excel}")
            self.write_log(f"Waterfall PNG: {ausgabe_wasserfall_png}")
            self.write_log(f"Waterfall PDF: {ausgabe_wasserfall_pdf}")
            self.write_log(f"Rotatable waterfall HTML: {ausgabe_wasserfall_html}")
            self.write_log(f"Control plot PNG: {ausgabe_kontrolle_png}")
            self.write_log(f"Baseline control PNG: {ausgabe_baseline_kontrolle_png}")

            messagebox.showinfo(
                "Finished",
                f"Raman analysis completed.\n\nOutput folder:\n{out_dir}"
            )

        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.write_log(f"ERROR: {e}")


def main():
    root = tk.Tk()
    app = RamanAuswertungGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
